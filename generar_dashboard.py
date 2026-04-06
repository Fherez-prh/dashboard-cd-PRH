import pandas as pd, base64, os, sys, openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

# ── LOGO ─────────────────────────────────────────────────────────────────────
with open(os.path.join(os.path.dirname(os.path.abspath(__file__)), 'html_template.txt'),'r',encoding='utf-8') as f:
    _tpl_raw = f.read()
import re as _re
_lm = _re.search(r'base64,([A-Za-z0-9+/=]+)', _tpl_raw)
LOGO = _lm.group(1) if _lm else ""

script_dir = os.path.dirname(os.path.abspath(__file__))

# ── ENCONTRAR EXCEL ───────────────────────────────────────────────────────────
excel_file = None
for fname in ['CAPACIDAD_.xlsx','CAPACIDAD_CON_INCIDENCIAS.xlsx']:
    p = os.path.join(script_dir, fname)
    if os.path.exists(p):
        try:
            xl = pd.ExcelFile(p)
            if 'BD' in xl.sheet_names:
                excel_file = p
                break
        except: pass

if not excel_file:
    for fname in os.listdir(script_dir):
        if fname.endswith('.xlsx') or fname.endswith('.xlsm'):
            p = os.path.join(script_dir, fname)
            try:
                xl = pd.ExcelFile(p)
                if 'BD' in xl.sheet_names and 'INCIDENCIA BARRIDO' in xl.sheet_names:
                    excel_file = p
                    break
            except: pass

if not excel_file:
    input("ERROR: No se encontro el Excel. Presiona Enter para salir.")
    sys.exit(1)

print(f"Leyendo: {os.path.basename(excel_file)}")

# ════════════════════════════════════════════════════════════════════════
# LEER BD — todo desde una sola hoja
# ════════════════════════════════════════════════════════════════════════
df = pd.read_excel(excel_file, sheet_name='BD', header=0)
df.columns = [str(c).strip() for c in df.columns]
df = df.drop(columns=[c for c in df.columns if 'Unnamed' in str(c)], errors='ignore')

# Detectar columnas clave
col_zona    = next((c for c in df.columns if c.upper()=='ZONA'), df.columns[0])
col_pasillo = next((c for c in df.columns if 'PASILLO' in c.upper()), df.columns[1])
col_nivel   = next((c for c in df.columns if 'NIVEL' in c.upper()), df.columns[3])
col_area    = next((c for c in df.columns if 'AREA' in c.upper() or 'CODIGO' in c.upper()), df.columns[6])
col_uni     = next((c for c in df.columns if 'UNIDAD' in c.upper()), df.columns[7])
col_status  = next((c for c in df.columns if 'STATUS' in c.upper()), df.columns[8])
col_ubi     = next((c for c in df.columns if 'UBICAC' in c.upper()), df.columns[5])
col_ibar    = next((c for c in df.columns if 'BARRIDO' in c.upper() and 'INCID' in c.upper()), None)
col_idup    = next((c for c in df.columns if 'DUPLIC' in c.upper() and 'INCID' in c.upper()), None)

df['_ZONA']   = df[col_zona].astype(str).str.strip()
df['_PAS']    = df[col_pasillo].astype(str).str.strip()
df['_NIVEL']  = pd.to_numeric(df[col_nivel], errors='coerce')
df['_AREA']   = df[col_area].astype(str).str.strip().str.upper()
df['_UNI']    = pd.to_numeric(df[col_uni], errors='coerce').fillna(0)
df['_STATUS'] = df[col_status].astype(str).str.strip().str.upper()
df['_UBI']    = df[col_ubi].astype(str).str.strip()

# Incidencias desde BD directamente
if col_ibar:
    incid_bar = df[df[col_ibar].astype(str).str.strip().str.upper().isin(['PALLET MAL UBICADO','SI'])][
        [col_ubi, col_ibar]].copy()
    incid_bar.columns = ['UBICACION','ESTADO']
    incid_bar = incid_bar.reset_index(drop=True)
else:
    incid_bar = pd.DataFrame(columns=['UBICACION','ESTADO'])

if col_idup:
    incid_dup = df[pd.to_numeric(df[col_idup], errors='coerce').fillna(0) > 0][
        [col_ubi, col_zona, col_idup]].copy()
    incid_dup.columns = ['UBICACION','ZONA','N_DUPLICADOS']
    incid_dup = incid_dup.reset_index(drop=True)
else:
    incid_dup = pd.DataFrame(columns=['UBICACION','ZONA','N_DUPLICADOS'])

print(f"  BD: {len(df):,} registros")
print(f"  Incidencias Barrido: {len(incid_bar)}")
print(f"  Incidencias Duplicados: {len(incid_dup)}")

# ════════════════════════════════════════════════════════════════════════
# AVERIADOS
# ════════════════════════════════════════════════════════════════════════
av_data = {}
# Search for averiados file
av_file = None
for fname in os.listdir(script_dir):
    if fname.endswith('.xlsx') and fname != os.path.basename(excel_file) and 'DESPACHO' not in fname.upper():
        p = os.path.join(script_dir, fname)
        try:
            xl2 = pd.ExcelFile(p)
            if 'datasap' in xl2.sheet_names or 'Datasap' in xl2.sheet_names:
                av_file = p
                break
        except: pass

if av_file:
    print(f"  Leyendo averiados: {os.path.basename(av_file)}")
    sh = 'datasap' if 'datasap' in pd.ExcelFile(av_file).sheet_names else 'Datasap'
    av = pd.read_excel(av_file, sheet_name=sh, header=0)
    av.columns = [str(c).strip() for c in av.columns]
    col_mat  = next((c for c in av.columns if 'MATERIAL' in c.upper()), av.columns[0])
    col_cen  = next((c for c in av.columns if 'CENTRO' in c.upper()), av.columns[1])
    col_alm  = next((c for c in av.columns if 'ALM' in c.upper()), av.columns[2])
    col_uni  = next((c for c in av.columns if 'LIBRE' in c.upper() or 'UNIDAD' in c.upper()), av.columns[3])
    col_desc = next((c for c in av.columns if 'DESCRIPCION' in c.upper()), av.columns[4])
    av['_UNI'] = pd.to_numeric(av[col_uni], errors='coerce').fillna(0)
    av['_ALM'] = av[col_alm].astype(str).str.strip()
    av['_CEN'] = av[col_cen].astype(str).str.strip()
    av['_DESC'] = av[col_desc].astype(str).str.strip()
    av['_MAT'] = av[col_mat].astype(str).str.strip()

    centros = sorted(av['_CEN'].unique().tolist())
    total_uni_av = int(av['_UNI'].sum())
    total_mat_av = int(av['_MAT'].nunique())

    # Read pallet data from Averiados sheet
    pal_map = {}  # {alm: pallet_count}
    total_pal_av = 0
    try:
        df_av_sheet = pd.read_excel(av_file, sheet_name='Averiados ', header=None)
        # Find pallet section (Cant. Pallet)
        for i, row in df_av_sheet.iterrows():
            if str(row.iloc[1]).strip() == 'Cant. Pallet':
                # Read rows below until Total general
                j = i + 1
                while j < len(df_av_sheet):
                    alm = str(df_av_sheet.iloc[j, 0]).strip()
                    pal = df_av_sheet.iloc[j, 1]
                    if alm == 'Total general':
                        total_pal_av = int(pal) if pd.notna(pal) else 0
                        break
                    if alm not in ['nan',''] and pd.notna(pal):
                        pal_map[alm] = int(pal)
                    j += 1
                break
    except: pass

    def av_stats(df_f):
        por_alm = df_f.groupby('_ALM').agg(UNI=('_UNI','sum'),MAT=('_MAT','nunique')).reset_index().sort_values('UNI',ascending=False)
        por_desc = df_f.groupby('_DESC').agg(UNI=('_UNI','sum')).reset_index().sort_values('UNI',ascending=False)
        alm_list = [{'a':str(r['_ALM']),'u':int(r['UNI']),'m':int(r['MAT']),'p':pal_map.get(str(r['_ALM']),0)} for _,r in por_alm.iterrows()]
        desc_list = [{'d':str(r['_DESC']),'u':int(r['UNI'])} for _,r in por_desc.iterrows()]
        total_p = sum(r['p'] for r in alm_list)
        return {'alm':alm_list,'desc':desc_list,'total_uni':int(df_f['_UNI'].sum()),'total_mat':int(df_f['_MAT'].nunique()),'total_pal':total_p}

    data_all = av_stats(av)
    data_centro = {}
    for cen in centros:
        data_centro[str(cen)] = av_stats(av[av['_CEN']==cen])

    def av_js_obj(d):
        alm_js = '['+','.join('{a:"'+r['a']+'"'+',u:'+str(r['u'])+',m:'+str(r['m'])+',p:'+str(r.get('p',0))+'}' for r in d['alm'])+']'
        desc_js = '['+','.join('{d:"'+r['d'].replace('"','').replace("'",'')+'"'+',u:'+str(r['u'])+'}' for r in d['desc'])+']'
        return '{alm:'+alm_js+',desc:'+desc_js+',total_uni:'+str(d['total_uni'])+',total_mat:'+str(d['total_mat'])+',total_pal:'+str(d.get('total_pal',0))+'}'


    centros_js = '['+','.join("'"+str(c)+"'" for c in centros)+']'
    data_centro_js = '{'+','.join("'"+k+"':"+av_js_obj(v) for k,v in data_centro.items())+'}'

    av_data = {
        'total_uni': total_uni_av,
        'total_mat': total_mat_av,
        'centros_js': centros_js,
        'data_all_js': av_js_obj(data_all),
        'data_centro_js': data_centro_js
    }
    print(f"  Averiados: {total_uni_av:,} unidades | {total_mat_av:,} materiales | {len(centros)} centros")
else:
    print("  Averiados: archivo SAP no encontrado en la carpeta")

# ════════════════════════════════════════════════════════════════════════
# CALCULAR KPIs CAPACIDAD
# ════════════════════════════════════════════════════════════════════════
ocu = df[df['_STATUS']=='OCUPADA']
vac = df[df['_STATUS']=='VACIA']
fds = df[df['_STATUS']=='FUERA DE SISTEMA']
total_uni = int(ocu['_UNI'].sum())
grand = len(df)

def zona_stats(z):
    zdf = df[df['_ZONA']==z]
    zo = zdf[zdf['_STATUS']=='OCUPADA']
    return len(zo),len(zdf[zdf['_STATUS']=='VACIA']),len(zdf[zdf['_STATUS']=='FUERA DE SISTEMA']),int(zo['_UNI'].sum())

def area_stats(key):
    a = df[df['_AREA'].str.contains(key,na=False)]
    ao = a[a['_STATUS']=='OCUPADA']
    return len(ao),len(a[a['_STATUS']=='VACIA']),len(a[a['_STATUS']=='FUERA DE SISTEMA']),int(ao['_UNI'].sum())

prs_o,prs_v,prs_f,prs_u = zona_stats('PRS')
pmz_o,pmz_v,pmz_f,pmz_u = zona_stats('PMZ')
pgr_o,pgr_v,pgr_f,pgr_u = zona_stats('PGR')
rs_o,rs_v,rs_f,rs_u   = area_stats('RACK SELECTIVO')
m1_o,m1_v,m1_f,m1_u   = area_stats('PRIMER PISO')
m2_o,m2_v,m2_f,m2_u   = area_stats('SEGUNDO PISO')
rg_o,rg_v,rg_f,rg_u   = area_stats('GRILLA')

nivs = []
prs_df = df[df['_ZONA']=='PRS']
for n in range(1,10):
    ndf = prs_df[prs_df['_NIVEL']==n]
    nivs.append((n,len(ndf[ndf['_STATUS']=='OCUPADA']),len(ndf[ndf['_STATUS']=='FUERA DE SISTEMA'])))

hm = {'PRS':[],'PMZ':[],'PGR':[]}
for pas,grp in df.groupby('_PAS'):
    z = grp['_ZONA'].iloc[0]
    if z not in hm: continue
    go = grp[grp['_STATUS']=='OCUPADA']
    hm[z].append({'p':str(pas),'o':len(go),'v':len(grp[grp['_STATUS']=='VACIA']),'f':len(grp[grp['_STATUS']=='FUERA DE SISTEMA']),'u':int(go['_UNI'].sum())})

# ════════════════════════════════════════════════════════════════════════
# CONSTRUIR JS
# ════════════════════════════════════════════════════════════════════════
def hmjs(lst):
    return '['+','.join('{p:\''+r['p']+'\',o:'+str(r['o'])+',v:'+str(r['v'])+',f:'+str(r['f'])+',u:'+str(r['u'])+'}' for r in lst)+']'

def bar_js(df_inc):
    parts = []
    for r in df_inc.itertuples(index=False):
        parts.append('{u:\''+str(r.UBICACION)+'\',e:\''+str(r.ESTADO)+'\'}')
    return '['+','.join(parts)+']'

def dup_js(df_inc):
    parts = []
    for r in df_inc.itertuples(index=False):
        parts.append('{u:\''+str(r.UBICACION)+'\',z:\''+str(r.ZONA)+'\',n:'+str(int(r.N_DUPLICADOS))+'}')
    return '['+','.join(parts)+']'

niv_js = ','.join('{n:\'N'+str(n)+'\',o:'+str(o)+',f:'+str(f)+'}' for n,o,f in nivs)

js = "const R={"
js += "global:{OCUPADA:"+str(len(ocu))+",VACIA:"+str(len(vac))+",FDS:"+str(len(fds))+",UNIDADES:"+str(total_uni)+"},"
js += "zonas:["
js += "{z:'PRS',o:"+str(prs_o)+",v:"+str(prs_v)+",f:"+str(prs_f)+",u:"+str(prs_u)+"},"
js += "{z:'PMZ',o:"+str(pmz_o)+",v:"+str(pmz_v)+",f:"+str(pmz_f)+",u:"+str(pmz_u)+"},"
js += "{z:'PGR',o:"+str(pgr_o)+",v:"+str(pgr_v)+",f:"+str(pgr_f)+",u:"+str(pgr_u)+"}],"
js += "areas:["
js += "{a:'Rack Selectivo',z:'PRS',o:"+str(rs_o)+",v:"+str(rs_v)+",f:"+str(rs_f)+",u:"+str(rs_u)+"},"
js += "{a:'Mezz. 1er Piso',z:'PMZ',o:"+str(m1_o)+",v:"+str(m1_v)+",f:"+str(m1_f)+",u:"+str(m1_u)+"},"
js += "{a:'Mezz. 2do Piso',z:'PMZ',o:"+str(m2_o)+",v:"+str(m2_v)+",f:"+str(m2_f)+",u:"+str(m2_u)+"},"
js += "{a:'Rack Grilla',z:'PGR',o:"+str(rg_o)+",v:"+str(rg_v)+",f:"+str(rg_f)+",u:"+str(rg_u)+"}],"
js += "niv:["+niv_js+"],"
js += "hm:{PRS:"+hmjs(hm['PRS'])+",PMZ:"+hmjs(hm['PMZ'])+",PGR:"+hmjs(hm['PGR'])+"},"
js += "incid_bar:"+str(len(incid_bar))+","
js += "incid_dup:"+str(len(incid_dup))+","
js += "incid_bar_data:"+bar_js(incid_bar)+","
js += "incid_dup_data:"+dup_js(incid_dup)+","

# Averiados
if av_data:
    js += "av:{"
    js += "total_uni:"+str(av_data['total_uni'])+","
    js += "total_mat:"+str(av_data['total_mat'])+","
    js += "centros:"+av_data['centros_js']+","
    js += "data_all:"+av_data['data_all_js']+","
    js += "data_centro:"+av_data['data_centro_js']
    js += "}"
else:
    js += "av:null"
js += "};"
# ── GENERAR HTML ──────────────────────────────────────────────────────────────
tpl_path = os.path.join(script_dir,'html_template.txt')
with open(tpl_path,'r',encoding='utf-8') as f:
    html = f.read()
html = html.replace('DATA_PLACEHOLDER', js)
out_html = os.path.join(script_dir,'index.html')
with open(out_html,'w',encoding='utf-8') as f:
    f.write(html)

# ── ACTUALIZAR HOJAS INCIDENCIAS EN EXCEL ─────────────────────────────────────
OR="E8581A"; GY="4A4A4A"; WH="FFFFFF"; LG="F2F2F2"; RD="C0392B"
def fill(c): return PatternFill("solid",fgColor=c)
def font(c=WH,sz=11,bold=False): return Font(name="Arial",size=sz,bold=bold,color=c)
def align(h="center",v="center"): return Alignment(horizontal=h,vertical=v)
def brd():
    s=Side(border_style="thin",color="DDDDDD")
    return Border(left=s,right=s,top=s,bottom=s)

try:
    wb = load_workbook(excel_file)
    for sname in ['INCID. BARRIDO','INCID. DB - PMZ P1']:
        if sname in wb.sheetnames: del wb[sname]

    # INCID. BARRIDO
    ws1 = wb.create_sheet('INCID. BARRIDO')
    ws1.sheet_properties.tabColor = RD
    ws1.sheet_view.showGridLines = False
    for ci,w in enumerate([3,30,22,3]): ws1.column_dimensions[get_column_letter(ci+1)].width=w
    for row in ws1.iter_rows(min_row=1,max_row=len(incid_bar)+8,min_col=1,max_col=5):
        for cell in row: cell.fill=fill(LG)
    ws1.row_dimensions[2].height=36
    ws1.merge_cells("B2:C2")
    c=ws1.cell(2,2,f"  ⚠  INCIDENCIAS BARRIDO ({datetime.now().strftime('%d/%m/%Y %H:%M')})  |  Total: {len(incid_bar)}")
    c.font=font(WH,12,True); c.fill=fill(RD); c.alignment=align("left","center")
    ws1.row_dimensions[4].height=22
    for ci,h in enumerate(['UBICACION','ESTADO']):
        c=ws1.cell(4,ci+2,h); c.font=font(WH,9,True); c.fill=fill(GY); c.alignment=align(); c.border=brd()
    if len(incid_bar)==0:
        ws1.merge_cells("B5:C5")
        c=ws1.cell(5,2,"  Sin incidencias de barrido"); c.font=font("15803D",10,True); c.fill=fill("F0FFF4"); c.alignment=align("left","center")
    else:
        for ri,row in enumerate(incid_bar.itertuples(index=False)):
            r=5+ri; ws1.row_dimensions[r].height=17
            bg=WH if ri%2==0 else "FEF2F2"
            for ci,val in enumerate([row.UBICACION,row.ESTADO]):
                c=ws1.cell(r,ci+2,val); c.font=font(GY,9); c.fill=fill(bg); c.alignment=align("left"); c.border=brd()
    ws1.auto_filter.ref="B4:C4"

    # INCID. DB - PMZ P1
    ws2 = wb.create_sheet('INCID. DB - PMZ P1')
    ws2.sheet_properties.tabColor = OR
    ws2.sheet_view.showGridLines = False
    for ci,w in enumerate([3,30,10,12,3]): ws2.column_dimensions[get_column_letter(ci+1)].width=w
    for row in ws2.iter_rows(min_row=1,max_row=len(incid_dup)+8,min_col=1,max_col=6):
        for cell in row: cell.fill=fill(LG)
    ws2.row_dimensions[2].height=36
    ws2.merge_cells("B2:D2")
    c=ws2.cell(2,2,f"  ⚠  INCIDENCIAS D.B - PMZ P1 ({datetime.now().strftime('%d/%m/%Y %H:%M')})  |  Total: {len(incid_dup)}")
    c.font=font(WH,12,True); c.fill=fill(OR); c.alignment=align("left","center")
    ws2.row_dimensions[4].height=22
    for ci,h in enumerate(['UBICACION','ZONA','N DUPLICADOS']):
        c=ws2.cell(4,ci+2,h); c.font=font(WH,9,True); c.fill=fill(GY); c.alignment=align(); c.border=brd()
    if len(incid_dup)==0:
        ws2.merge_cells("B5:D5")
        c=ws2.cell(5,2,"  Sin duplicados en PMZ Piso 1"); c.font=font("15803D",10,True); c.fill=fill("F0FFF4"); c.alignment=align("left","center")
    else:
        for ri,row in enumerate(incid_dup.itertuples(index=False)):
            r=5+ri; ws2.row_dimensions[r].height=17
            bg=WH if ri%2==0 else "FFF5EC"
            for ci,val in enumerate([row.UBICACION,row.ZONA,int(row.N_DUPLICADOS)]):
                c=ws2.cell(r,ci+2,val); c.font=font(GY,9); c.fill=fill(bg); c.alignment=align("left" if ci<2 else "center"); c.border=brd()
    ws2.auto_filter.ref="B4:D4"

    # Intentar guardar - puede fallar si el Excel está abierto
    import tempfile, shutil
    tmp = excel_file + '.tmp_dashboard'
    wb.save(tmp)
    try:
        shutil.move(tmp, excel_file)
        print("  Hojas incidencias actualizadas OK")
    except Exception:
        os.remove(tmp)
        print("  Aviso: cierra el Excel para actualizar las hojas de incidencias")
except Exception as e:
    print(f"  Aviso hojas incidencias: {e} (cierra el Excel y vuelve a correr para actualizar hojas)")

# ── RESULTADO ─────────────────────────────────────────────────────────────────
pct = len(ocu)/grand*100 if grand else 0
print()
print("="*55)
print("  TODO GENERADO EXITOSAMENTE!")
print("="*55)
print(f"  Dashboard:          index.html")
print(f"  Total ubicaciones:  {grand:,}")
print(f"  Ocupacion:          {pct:.1f}%")
print(f"  Unidades stock:     {total_uni:,}")
print(f"  Incid. Barrido:     {len(incid_bar)}")
print(f"  Incid. Duplicados:  {len(incid_dup)}")

print(f"  Fecha:              {datetime.now().strftime('%d/%m/%Y %H:%M')}")
print()
print("  SIGUIENTE PASO: Sube index.html a GitHub Pages")
print("="*55)
input("\nPresiona Enter para cerrar...")
